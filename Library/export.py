from fpdf import FPDF
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from Library import group_sheet, get_sheet
import json
import urllib.parse
import numpy as np

# Função para carregar as informações dos caminhões do arquivo JSON
def load_truck_data(filename='./Data/caminhões.json'):
    try:
        with open(filename, 'r', encoding='utf-8') as file:
            truck_data = json.load(file)
            # Cria um dicionário com o número do caminhão como chave e o volume como valor
            truck_size = {truck['TAMANHO CAMINHÃO ']: truck['volume'] for truck in truck_data}
            return truck_size
    except Exception as e:
        print(f"Erro ao carregar o arquivo {filename}: {e}")
        return {}

def create_pdf(routes, truck_number, truck_size, output_file='roteiro_entregas.pdf'):
    # Cria o PDF com as rotas
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)

    def safe_value(value):
        """Trata valores nulos ou NaN para evitar erros."""
        if pd.isna(value) or value is None:
            return 'Sem informações'
        return str(value)

    pdf.set_font("Arial", size=12)

    # Obtém o volume do caminhão com base no número
    truck_key = f"CAMINHÃO {truck_number}"  # Exemplo: "CAMINHÃO 1"

    # Procurando o caminhão correto no dicionário de tamanhos de caminhões
    truck_data = truck_size.get(truck_key, None)

    if truck_data:
        truck_volume = truck_data  # O volume é diretamente o valor da chave
    else:
        truck_volume = "Desconhecido"  # Se o caminhão não for encontrado, atribui "Desconhecido"

    for date, daily_route in routes.items():
        pdf.add_page()
        pdf.set_font("Arial", style='B', size=12)
        pdf.cell(200, 10, txt=f"Data de Entrega: {date}", ln=True)
        pdf.ln(5)  # Linha em branco

        # Dados do caminhão e equipe
        equipe = f"Equipe {truck_number}"
        total_clientes = sum(len(clients) for clients in daily_route.values())

        # A lógica para determinar o tamanho do caminhão com base no volume
        if truck_volume != "Desconhecido" and float(truck_volume.replace('m³', '').replace(',', '.')) > 20:
            tamanho_caminhao = "Grande"
        else:
            tamanho_caminhao = "Pequeno"

        pdf.set_font("Arial", style='B', size=10)
        pdf.cell(200, 10, txt=f"Caminhão {truck_number}", ln=True)
        pdf.cell(200, 10, txt=f"Equipe: {equipe}", ln=True)
        pdf.cell(200, 10, txt=f"Total de Clientes: {total_clientes}", ln=True)
        pdf.cell(200, 10, txt=f"Tamanho do Caminhão: {tamanho_caminhao} (Volume: {truck_volume})", ln=True)
        pdf.ln(5)  # Linha em branco

        for session, clients in daily_route.items():
            pdf.set_font("Arial", style='B', size=10)
            pdf.cell(200, 10, txt=f"\n{session.capitalize()}:", ln=True)
            pdf.set_font("Arial", size=10)

            for idx, client in enumerate(clients, start=1):
                if not isinstance(client, dict):
                    continue  # Pula itens que não sejam dicionários

                pdf.cell(200, 10, txt=f"Parada {idx}", ln=True)

                # Validação e conversão de 'Medida Sofá'
                sofa_measure = client.get('Medida Sofá', '').strip()
                try:
                    if sofa_measure.isdigit():  # Verifica se é numérico
                        sofa_measure = round(float(sofa_measure) * 0.1, 2)
                    else:
                        sofa_measure = "Valor inválido"
                except ValueError:
                    sofa_measure = "Valor inválido"

                client_info = (
                    f"Nº do pedido: {safe_value(client.get('N\u00b0 de pedido'))} | "
                    f"Data de entrega: {safe_value(client.get('Data Entrega'))} | "
                    f"CEP: {safe_value(client.get('Cep'))} | "
                    f"Número da casa: {safe_value(client.get('N\u00b0 Casa/Ap'))} | "
                    f"Endereço: {safe_value(client.get('endereco'))} | "
                    f"Modelo do Sofá: {safe_value(client.get('Modelo Sofá'))} | "
                    f"Medida Sofá: {safe_value(sofa_measure)} | "
                    f"Características Sofá: {safe_value(client.get('Características sofá'))} | "
                    f"Local: {safe_value(client.get('Local'))} | "
                    f"Andar: {safe_value(client.get('andar'))} | "
                    f"Restrição de local: {safe_value(client.get('Restrições local Elevador'))} | "
                    f"Restrição de Horário: {safe_value(client.get('restrições Horário'))} | "
                )

                pdf.multi_cell(0, 10, txt=client_info)
                pdf.ln(2)  # Linha em branco para separar as paradas

    pdf.output(output_file)
    print(f"PDF gerado com sucesso em {output_file}")

def create_xlsx(routes, output_file='roteiro_entregas.xlsx'):
    data_for_export = []

    def safe_value(value):
        if pd.isna(value) or value is None:
            return 'Sem informações'
        return str(value)

    for date, daily_route in routes.items():
        for session, clients in daily_route.items():
            for client in clients:
                if isinstance(client, dict):  # Verifica se 'client' é um dicionário
                    if "retorno ao ponto de partida" not in str(client.get('status')).lower():
                        data_for_export.append({
                            "Nº Pedido": safe_value(client.get('N\u00b0 de pedido')),
                            "Data de entrega": safe_value(client.get('Data Entrega')),
                            "Período": session,
                            "Nome do cliente": safe_value(client.get('Nome cliente', 'Desconhecido')),
                            "Bairro": safe_value(client.get('bairro')),
                            "Modelo do sofá": safe_value(client.get('Modelo Sofá', 'Desconhecido'))
                        })

    # Converte os dados em um DataFrame e salva no arquivo Excel
    df = pd.DataFrame(data_for_export)
    df.to_excel(output_file, index=False, engine='openpyxl')

    # Manipulação do estilo da planilha
    wb = load_workbook(output_file)
    ws = wb.active

    bold_font = Font(bold=True)
    center_alignment = Alignment(horizontal='center', vertical='center')

    for cell in ws[1]:
        cell.font = bold_font
        cell.alignment = center_alignment

    # Ajusta as larguras das colunas
    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col)  # Melhor forma de calcular a largura
        adjusted_width = (max_length + 2)
        ws.column_dimensions[col[0].column_letter].width = adjusted_width

    # Alinha as células da planilha
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = center_alignment

    wb.save(output_file)
    print(f"Arquivo XLSX gerado com sucesso em {output_file}")

def create_links_txt(deliveries_json, output_file="route_link.txt"):
    try:
        base_url = "https://www.google.com/maps/dir/"
        addresses = []

        # Iterar por datas e turnos no JSON
        for date, shifts in deliveries_json.items():
            for shift, deliveries in shifts.items():
                for delivery in deliveries:
                    if "endereco" in delivery:
                        addresses.append(delivery["endereco"])

        # Escapar os endereços para o formato URL
        encoded_addresses = [urllib.parse.quote_plus(address) for address in addresses]

        # Criar o link concatenando os endereços
        route_link = base_url + "/".join(encoded_addresses)

        # Salvar o link em um arquivo texto
        with open(output_file, "w") as file:
            file.write(route_link)
        
        print(f"Link salvo no arquivo: {output_file}")
        return route_link
    except Exception as e:
        print(f"Erro ao exportar a rota: {e}")